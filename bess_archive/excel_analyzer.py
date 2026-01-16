import pandas as pd
import openpyxl
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

def analyze_excel_structure(file_path):
    """Analyze the Excel file structure and extract key information"""
    print(f"Analyzing: {file_path}")
    print(f"File size: {pd.read_excel(file_path, engine='openpyxl').memory_usage(deep=True).sum() / 1024 / 1024:.2f} MB")
    
    # Load workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    
    print(f"\nWorksheet Names: {wb.sheetnames}")
    print(f"Total worksheets: {len(wb.sheetnames)}")
    
    # Analyze each worksheet
    worksheet_info = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Get dimensions
        max_row = ws.max_row
        max_col = ws.max_column
        
        # Count non-empty cells
        non_empty_cells = 0
        formulas = 0
        
        for row in ws.iter_rows(max_row=max_row, max_col=max_col):
            for cell in row:
                if cell.value is not None:
                    non_empty_cells += 1
                if cell.data_type == 'f':  # Formula
                    formulas += 1
        
        worksheet_info[sheet_name] = {
            'dimensions': f"{max_row} rows x {max_col} cols",
            'non_empty_cells': non_empty_cells,
            'formulas': formulas,
            'density': non_empty_cells / (max_row * max_col) * 100
        }
    
    return worksheet_info, wb.sheetnames

def extract_sample_data(file_path, sheet_names, max_rows=10):
    """Extract sample data from key worksheets"""
    sample_data = {}
    
    for sheet_name in sheet_names:
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=max_rows, engine='openpyxl')
            sample_data[sheet_name] = df
            print(f"\n--- {sheet_name} ---")
            print(f"Shape: {df.shape}")
            print("Columns:", list(df.columns))
            print("Sample data:")
            print(df.head())
        except Exception as e:
            print(f"Error reading {sheet_name}: {e}")
    
    return sample_data

if __name__ == "__main__":
    file_path = "AUDIT 20251201 40MW Solar ^M BESS Ecoplexus.xlsx"
    
    try:
        # Analyze structure
        worksheet_info, sheet_names = analyze_excel_structure(file_path)
        
        print("\n=== WORKSHEET ANALYSIS ===")
        for sheet, info in worksheet_info.items():
            print(f"\n{sheet}:")
            print(f"  Dimensions: {info['dimensions']}")
            print(f"  Non-empty cells: {info['non_empty_cells']}")
            print(f"  Formulas: {info['formulas']}")
            print(f"  Data density: {info['density']:.1f}%")
        
        # Extract sample data from first few sheets
        key_sheets = sheet_names[:5]  # First 5 sheets
        sample_data = extract_sample_data(file_path, key_sheets)
        
    except Exception as e:
        print(f"Error analyzing Excel file: {e}")
