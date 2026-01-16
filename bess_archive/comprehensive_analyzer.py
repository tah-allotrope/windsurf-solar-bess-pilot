import pandas as pd
import numpy as np
from openpyxl import load_workbook
import warnings
warnings.filterwarnings('ignore')

def comprehensive_sheet_analysis():
    """Analyze ALL sheets in the Excel file comprehensively"""
    file_path = "AUDIT 20251201 40MW Solar ^M BESS Ecoplexus.xlsx"
    
    # Get all sheet names first
    wb = load_workbook(file_path, read_only=True)
    all_sheets = wb.sheetnames
    wb.close()
    
    print(f"Total sheets found: {len(all_sheets)}")
    print(f"Sheet names: {all_sheets}")
    print("="*80)
    
    # Analyze each sheet in detail
    sheet_analysis = {}
    
    for sheet_name in all_sheets:
        print(f"\n{'='*20} ANALYZING: {sheet_name} {'='*20}")
        
        try:
            # Read the sheet
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            # Basic info
            print(f"Dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
            
            # Find actual data range (skip empty borders)
            non_empty_rows = []
            for i, row in df.iterrows():
                if not row.isna().all():
                    non_empty_rows.append(i)
            
            if non_empty_rows:
                first_data_row = min(non_empty_rows)
                last_data_row = max(non_empty_rows)
                print(f"Data range: Row {first_data_row} to {last_data_row}")
                
                # Sample the first few data rows
                sample_data = df.iloc[first_data_row:min(first_data_row+10, last_data_row+1)]
                print("Sample data:")
                print(sample_data.to_string())
                
                # Look for headers (non-empty first row)
                first_row = df.iloc[first_data_row]
                if not first_row.isna().all():
                    print(f"\nPotential headers found:")
                    for i, val in enumerate(first_row):
                        if pd.notna(val) and str(val).strip():
                            print(f"  Column {i}: {val}")
                
                # Try to detect if this sheet has structured data
                if df.shape[0] > 50:  # Large dataset
                    print(f"\n*** LARGE DATASET DETECTED ***")
                    # Try reading with first row as header
                    try:
                        df_with_header = pd.read_excel(file_path, sheet_name=sheet_name, nrows=5)
                        print("Columns with header interpretation:")
                        print(list(df_with_header.columns))
                    except:
                        pass
                
            else:
                print("No data found in this sheet")
            
            sheet_analysis[sheet_name] = {
                'shape': df.shape,
                'data_rows': len(non_empty_rows),
                'has_data': len(non_empty_rows) > 0
            }
            
        except Exception as e:
            print(f"Error reading {sheet_name}: {e}")
            sheet_analysis[sheet_name] = {'error': str(e)}
        
        print("-"*60)
    
    return sheet_analysis, all_sheets

def extract_key_sheets_data():
    """Extract data from key sheets that likely contain important model inputs"""
    file_path = "AUDIT 20251201 40MW Solar ^M BESS Ecoplexus.xlsx"
    
    key_sheets = ['Assumption', 'Financial', 'Measures', 'Lifetime', 'Helper', 'Other Input']
    extracted_data = {}
    
    for sheet_name in key_sheets:
        try:
            print(f"\n--- Extracting from {sheet_name} ---")
            
            # Read entire sheet
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            # Find data rows
            data_rows = []
            for i, row in df.iterrows():
                if not row.isna().all():
                    data_rows.append(row.tolist())
            
            extracted_data[sheet_name] = data_rows
            print(f"Extracted {len(data_rows)} data rows")
            
            # Print first few rows to understand structure
            for i, row in enumerate(data_rows[:10]):
                clean_row = [str(x) if pd.notna(x) else '' for x in row]
                if any(clean_row):  # Only show non-empty rows
                    print(f"Row {i}: {clean_row[:5]}...")  # Show first 5 columns
            
        except Exception as e:
            print(f"Error extracting from {sheet_name}: {e}")
    
    return extracted_data

if __name__ == "__main__":
    print("COMPREHENSIVE EXCEL MODEL ANALYSIS")
    print("="*80)
    
    # Analyze all sheets
    sheet_analysis, all_sheets = comprehensive_sheet_analysis()
    
    print(f"\n{'='*80}")
    print("SUMMARY OF ALL SHEETS")
    print("="*80)
    
    for sheet_name, info in sheet_analysis.items():
        if 'error' not in info:
            print(f"{sheet_name}: {info['shape'][0]}×{info['shape'][1]}, {info['data_rows']} data rows")
        else:
            print(f"{sheet_name}: ERROR - {info['error']}")
    
    print(f"\n{'='*80}")
    print("DETAILED EXTRACTION FROM KEY SHEETS")
    print("="*80)
    
    # Extract detailed data from key sheets
    key_data = extract_key_sheets_data()
